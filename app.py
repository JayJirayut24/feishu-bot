import os
import json
import re
import time
import unicodedata
import requests
import threading
from flask import Flask, request, jsonify
from io import BytesIO
from datetime import datetime, timezone, timedelta
from openpyxl import load_workbook

app = Flask(__name__)

# ---------------------------------------------------------
# ป้องกัน Feishu ส่ง Event ซ้ำ (Deduplication)
# ---------------------------------------------------------
processed_event_ids = set()
processed_event_lock = threading.Lock()

# ---------------------------------------------------------
# ตั้งค่า App ID และ App Secret ของ Feishu
# ---------------------------------------------------------
APP_ID = "cli_aabfaea8b0619bfc"
APP_SECRET = "3emUt5KWwH01BlhIKADP2bCb5C062oxt"

# ---------------------------------------------------------
# ตั้งค่า Feishu Wiki Spreadsheet ปลายทาง
# ---------------------------------------------------------
WIKI_TOKEN = "UbCZwapNyiN15YkEKADcFyUHnWf"
TARGET_SHEET_NAMES = ["ยิงส่ง - ITCBI", "ยิงถึง - ITCBI"]
BRANCH_CODE_SHEET = "ยิงส่ง - ITCBI"
SUMMARY_LOG_SHEET = "สรุปยอดยิงส่ง"

# ---------------------------------------------------------
# Bangkok Timezone (UTC+7)
# ---------------------------------------------------------
BKK_TZ = timezone(timedelta(hours=7))

# ---------------------------------------------------------
# Cache สำหรับ API token และ Sheet IDs (ลด API call ซ้ำ)
# ---------------------------------------------------------
_api_cache = {
    "token": None,
    "token_expires": 0.0,
    "stoken": None,
    "sheet_ids": None,
    "summary_sheet_id": None,
}
_api_cache_lock = threading.Lock()

# ---------------------------------------------------------
# ระบบเก็บ State จำนวน AWB แบบรายวัน และ รหัสสาขา
# ---------------------------------------------------------
STATE_FILE = "state.json"
app_state = {
    "current_date": "",
    "daily_count": 0,
    "branch_summary": {}
}

# โหลดข้อมูลเก่าถ้ามี
if os.path.exists(STATE_FILE):
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            app_state = json.load(f)
    except:
        pass

def save_state():
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(app_state, f, ensure_ascii=False)
    except:
        pass

def add_to_daily_count(new_count, branch_list):
    tz = timezone(timedelta(hours=7))
    now = datetime.now(tz)
    today_str = now.strftime("%d/%m/%Y")
    
    # ถ้าระบบข้ามวันแล้ว ให้รีเซ็ตตัวนับและสรุปยอดเป็น 0
    if app_state.get("current_date") != today_str:
        app_state["current_date"] = today_str
        app_state["daily_count"] = 0
        app_state["branch_summary"] = {}
        
    app_state["daily_count"] += new_count
    
    # นับยอดแต่ละรหัสสาขา
    for br in branch_list:
        if br:
            app_state["branch_summary"][br] = app_state["branch_summary"].get(br, 0) + 1
            
    save_state()
    return app_state["daily_count"], today_str


# ---------------------------------------------------------
# ฟังก์ชันขอ Token จาก Feishu
# ---------------------------------------------------------
def get_tenant_access_token():
    """ขอ Tenant Access Token จาก Feishu (cache ไว้ตลอด expire)"""
    with _api_cache_lock:
        if _api_cache["token"] and time.time() < _api_cache["token_expires"]:
            return _api_cache["token"]
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    res = requests.post(url, json={"app_id": APP_ID, "app_secret": APP_SECRET}).json()
    token = res.get("tenant_access_token")
    expire = int(res.get("expire", 7200))
    with _api_cache_lock:
        _api_cache["token"] = token
        _api_cache["token_expires"] = time.time() + expire - 120
    return token


# ---------------------------------------------------------
# ฟังก์ชันดาวน์โหลดไฟล์จากแชท Feishu
# ---------------------------------------------------------
def download_feishu_file(message_id, file_key, token):
    """ดาวน์โหลดไฟล์ที่ผู้ใช้ส่งมาในแชท โดยใช้ message_id และ file_key"""
    url = f"https://open.feishu.cn/open-apis/im/v1/messages/{message_id}/resources/{file_key}?type=file"
    headers = {"Authorization": f"Bearer {token}"}
    res = requests.get(url, headers=headers)
    if res.status_code == 200:
        return res.content
    return None


# ---------------------------------------------------------
# ฟังก์ชันดึงข้อมูลจากไฟล์ Excel
# ---------------------------------------------------------
def _cell_to_str(cell):
    if cell is None:
        return ""
    if isinstance(cell, float) and cell == int(cell):
        return str(int(cell))
    return str(cell).strip()

def _normalize(text):
    """Normalize Unicode + ลบ whitespace เพื่อเปรียบเทียบ header ภาษาไทยได้แม่นยำ"""
    return unicodedata.normalize("NFC", "".join(text.split()))

def extract_data_from_excel(file_bytes):
    try:
        wb = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
        awb_list = []
        branch_codes = []

        # ใช้เฉพาะ Sheet ที่ active ตอนที่ user save ไฟล์ (sheet หน้าล่าสุด)
        ws = wb.active
        if ws is None:
            wb.close()
            return [], []

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            wb.close()
            return [], []

        # === ค้นหา header "สาขาที่ถูกต้อง" ใน 5 แถวแรก ===
        TARGET = _normalize("สาขาที่ถูกต้อง")
        correct_branch_col = None
        awb_col = None
        data_start_row = 0

        for row_idx, row in enumerate(all_rows[:5]):
            if not row:
                continue
            for i, cell in enumerate(row):
                h = _normalize(_cell_to_str(cell))
                if TARGET in h:
                    correct_branch_col = i
                if h.upper() == "AWB":
                    awb_col = i
            if correct_branch_col is not None:
                data_start_row = row_idx + 1
                break

        if correct_branch_col is not None:
            for row in all_rows[data_start_row:]:
                if not row:
                    continue

                awb_str = ""
                if awb_col is not None and awb_col < len(row):
                    candidate = _cell_to_str(row[awb_col])
                    if re.match(r'^\d{12}$', candidate) or re.match(r'^B\d+$', candidate):
                        awb_str = candidate

                if not awb_str:
                    for j, cell in enumerate(row):
                        if j == correct_branch_col:
                            continue
                        candidate = _cell_to_str(cell)
                        if re.match(r'^\d{12}$', candidate) or re.match(r'^B\d+$', candidate):
                            awb_str = candidate
                            break

                if not awb_str:
                    continue

                branch_str = ""
                if correct_branch_col < len(row):
                    candidate = _cell_to_str(row[correct_branch_col])
                    if re.match(r'^\d{6}$', candidate):
                        branch_str = candidate

                awb_list.append(awb_str)
                branch_codes.append(branch_str)

            # ลบ AWB ซ้ำ (พร้อม branch ที่จับคู่)
            seen = set()
            deduped_awb, deduped_br = [], []
            for awb, br in zip(awb_list, branch_codes):
                if awb not in seen:
                    seen.add(awb)
                    deduped_awb.append(awb)
                    deduped_br.append(br)
            awb_list, branch_codes = deduped_awb, deduped_br

        else:
            start_idx = 0
            if all_rows:
                first_row_cells = [_cell_to_str(c) for c in (all_rows[0] or []) if c is not None]
                has_data = any(
                    re.match(r'^\d{12}$', v) or re.match(r'^B\d+$', v) or re.match(r'^\d{6}$', v)
                    for v in first_row_cells
                )
                if not has_data:
                    start_idx = 1

            col_six = {}
            for row in all_rows[start_idx:]:
                if not row:
                    continue
                for j, cell in enumerate(row):
                    v = _cell_to_str(cell)
                    if v and re.match(r'^\d{6}$', v):
                        col_six.setdefault(j, []).append(v)

            # คอลัมน์ที่ทุกค่าเหมือนกัน (เช่น 812087 ทั้งหมด) = สาขาผิด → ข้าม
            skip_cols = {j for j, vals in col_six.items()
                         if len(vals) > 1 and len(set(vals)) == 1}

            for row in all_rows[start_idx:]:
                if not row:
                    continue
                for j, cell in enumerate(row):
                    if j in skip_cols:
                        continue
                    cell_str = _cell_to_str(cell)
                    if not cell_str:
                        continue
                    if re.match(r'^\d{12}$', cell_str):
                        awb_list.append(cell_str)
                    elif re.match(r'^B\d+$', cell_str):
                        awb_list.append(cell_str)
                    elif re.match(r'^\d{6}$', cell_str):
                        branch_codes.append(cell_str)

            seen = set()
            unique_awb = []
            for awb in awb_list:
                if awb not in seen:
                    seen.add(awb)
                    unique_awb.append(awb)
            awb_list = unique_awb

        wb.close()
        return awb_list, branch_codes

    except Exception:
        return [], []


# ---------------------------------------------------------
# ฟังก์ชันดึง Spreadsheet Token จาก Wiki Token
# ---------------------------------------------------------
def get_spreadsheet_token(token):
    """ดึง obj_token (Spreadsheet Token จริง) จาก Wiki Node (cached)"""
    with _api_cache_lock:
        if _api_cache["stoken"]:
            return _api_cache["stoken"], None
    url = f"https://open.feishu.cn/open-apis/wiki/v2/spaces/get_node?token={WIKI_TOKEN}"
    headers = {"Authorization": f"Bearer {token}"}
    try:
        res = requests.get(url, headers=headers).json()
        if res.get("code") != 0:
            return None, f"API Error: {res.get('msg')} (Code: {res.get('code')})"
        node = res.get("data", {}).get("node", {})
        stoken = node.get("obj_token")
        if stoken:
            with _api_cache_lock:
                _api_cache["stoken"] = stoken
        return stoken, None
    except Exception as e:
        return None, f"Request Error: {str(e)}"


# ---------------------------------------------------------
# ฟังก์ชันดึง Sheet ID ของ Sheet ที่ต้องการ
# ---------------------------------------------------------
def get_sheet_ids(spreadsheet_token, token):
    """ค้นหา Sheet ID ของ Sheet ชื่อ 'ยิงส่ง - ITCBI' และ 'ยิงถึง - ITCBI' (cached)"""
    with _api_cache_lock:
        if _api_cache["sheet_ids"]:
            return _api_cache["sheet_ids"]
    url = f"https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/{spreadsheet_token}/sheets/query"
    headers = {"Authorization": f"Bearer {token}"}
    try:
        res = requests.get(url, headers=headers).json()
        sheets = res.get("data", {}).get("sheets", [])

        sheet_ids = {}
        for sheet in sheets:
            title = sheet.get("title", "")
            sheet_id = sheet.get("sheet_id", "")
            if title in TARGET_SHEET_NAMES:
                sheet_ids[title] = sheet_id

        if sheet_ids:
            with _api_cache_lock:
                _api_cache["sheet_ids"] = sheet_ids
        return sheet_ids
    except Exception:
        return {}


# ---------------------------------------------------------
# ฟังก์ชันดึง Sheet ID ของ สรุปยอดยิงส่ง
# ---------------------------------------------------------
def get_summary_sheet_id(spreadsheet_token, token):
    url = f"https://open.feishu.cn/open-apis/sheets/v3/spreadsheets/{spreadsheet_token}/sheets/query"
    headers = {"Authorization": f"Bearer {token}"}
    try:
        res = requests.get(url, headers=headers).json()
        for sheet in res.get("data", {}).get("sheets", []):
            if sheet.get("title", "") == SUMMARY_LOG_SHEET:
                return sheet.get("sheet_id", "")
    except Exception:
        pass
    return None


# ---------------------------------------------------------
# บันทึก/อัปเดต log รายวันลง Sheet สรุปยอดยิงส่ง (upsert)
# ---------------------------------------------------------
def save_daily_log(date_str, branch_summary, spreadsheet_token, summary_sheet_id, token):
    if not branch_summary or not summary_sheet_id:
        return False

    api_headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    # 1. อ่านข้อมูลที่มีอยู่ใน log sheet
    read_url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{spreadsheet_token}/values/{summary_sheet_id}!A2:C5000"
    )
    try:
        raw = (
            requests.get(read_url, headers={"Authorization": f"Bearer {token}"})
            .json()
            .get("data", {}).get("valueRange", {}).get("values") or []
        )
    except Exception:
        raw = []

    # 2. หาแถวที่มี date_str อยู่แล้ว (sheet row number, 1-based)
    today_row_indices = []
    for i, row in enumerate(raw):
        if row and row[0] and str(row[0]).strip() == date_str:
            today_row_indices.append(i + 2)

    # 3. ลบแถวของวันนี้ก่อน (ถ้ามี) เพื่อ upsert
    if today_row_indices:
        delete_url = (
            f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
            f"{spreadsheet_token}/dimension_range"
        )
        payload = {
            "dimension": {
                "sheetId": summary_sheet_id,
                "majorDimension": "ROWS",
                "startIndex": min(today_row_indices),
                "endIndex": max(today_row_indices) + 1
            }
        }
        try:
            requests.delete(delete_url, headers=api_headers, json=payload)
        except Exception:
            pass

    # 4. Append แถวใหม่สำหรับแต่ละสาขา
    rows_to_append = [
        [date_str, br, count]
        for br, count in sorted(branch_summary.items())
    ]
    append_url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{spreadsheet_token}/values_append"
    )
    payload = {
        "valueRange": {
            "range": f"{summary_sheet_id}!A:C",
            "values": rows_to_append
        }
    }
    try:
        res = requests.post(
            append_url, headers=api_headers, json=payload,
            params={"insertDataOption": "INSERT_ROWS"}
        ).json()
        return res.get("code") == 0
    except Exception:
        return False


# ---------------------------------------------------------
# ลบแถวในชีตที่บันทึกมาจาก filename ที่ระบุ
# ---------------------------------------------------------
def delete_rows_by_filename(spreadsheet_token, sheet_id, filename_col_idx, filename, token):
    """ลบแถวที่ column filename_col_idx ตรงกับ filename (case-insensitive)
    คืนค่า (จำนวนที่ลบ, error_msg or None)
    """
    api_headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}

    end_col_letter = chr(ord('A') + filename_col_idx)
    read_url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{spreadsheet_token}/values/{sheet_id}!A2:{end_col_letter}5000"
    )
    try:
        raw = (
            requests.get(read_url, headers={"Authorization": f"Bearer {token}"})
            .json()
            .get("data", {}).get("valueRange", {}).get("values") or []
        )
    except Exception as e:
        return 0, str(e)

    fname_lower = filename.lower().strip()
    match_rows = []
    for i, row in enumerate(raw):
        if not row or not row[0] or str(row[0]).strip() in ("", "None"):
            continue
        cell_val = (
            str(row[filename_col_idx]).strip().lower()
            if len(row) > filename_col_idx and row[filename_col_idx]
            else ""
        )
        if cell_val and (fname_lower in cell_val or cell_val in fname_lower):
            match_rows.append(i + 2)  # 1-based sheet row

    if not match_rows:
        return 0, None

    count = len(match_rows)

    # จัดกลุ่ม contiguous ranges แล้วลบจากล่างขึ้นบน (ไม่ให้ index เลื่อน)
    sorted_desc = sorted(match_rows, reverse=True)
    ranges = []
    g_end = sorted_desc[0]
    g_start = sorted_desc[0]
    for r in sorted_desc[1:]:
        if r == g_start - 1:
            g_start = r
        else:
            ranges.append((g_start, g_end))
            g_end = r
            g_start = r
    ranges.append((g_start, g_end))

    delete_url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{spreadsheet_token}/dimension_range"
    )
    for (s, e) in ranges:
        try:
            requests.delete(delete_url, headers=api_headers, json={
                "dimension": {
                    "sheetId": sheet_id,
                    "majorDimension": "ROWS",
                    "startIndex": s,
                    "endIndex": e + 1
                }
            })
        except Exception:
            pass

    return count, None


# ---------------------------------------------------------
# ตัดยอดอัตโนมัติเวลา 00:00 น.
# ---------------------------------------------------------
def midnight_daily_cutoff():
    tz = timezone(timedelta(hours=7))
    yesterday_str = (datetime.now(tz) - timedelta(days=1)).strftime("%d/%m/%Y")
    try:
        token = get_tenant_access_token()
        if not token:
            return
        stoken, _ = get_spreadsheet_token(token)
        if not stoken:
            return
        sids = get_sheet_ids(stoken, token)
        song_id = sids.get(BRANCH_CODE_SHEET)
        if not song_id:
            return

        read_url = (
            f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
            f"{stoken}/values/{song_id}!A2:H5000"
        )
        raw = (
            requests.get(read_url, headers={"Authorization": f"Bearer {token}"})
            .json()
            .get("data", {}).get("valueRange", {}).get("values") or []
        )
        branch_summary = {}
        for row in raw:
            if not row or not row[0] or str(row[0]).strip() in ("", "None"):
                continue
            time_val = str(row[7]).strip() if len(row) > 7 and row[7] else ""
            if yesterday_str not in time_val:
                continue
            br = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            if br and br not in ("", "None"):
                branch_summary[br] = branch_summary.get(br, 0) + 1

        if branch_summary:
            sid = get_summary_sheet_id(stoken, token)
            if sid:
                save_daily_log(yesterday_str, branch_summary, stoken, sid, token)
    except Exception:
        pass


def start_midnight_scheduler():
    """Background thread ตัดยอดอัตโนมัติทุกเที่ยงคืน (Bangkok time)"""
    def _loop():
        while True:
            tz = timezone(timedelta(hours=7))
            now = datetime.now(tz)
            next_midnight = (now + timedelta(days=1)).replace(
                hour=0, minute=0, second=5, microsecond=0
            )
            time.sleep(max((next_midnight - now).total_seconds(), 1))
            try:
                midnight_daily_cutoff()
            except Exception:
                pass

    threading.Thread(target=_loop, daemon=True).start()


# ---------------------------------------------------------
# ฟังก์ชันเพิ่มข้อมูลลงใน Feishu Spreadsheet
# ---------------------------------------------------------
def append_to_feishu_sheet(spreadsheet_token, sheet_id, values, start_col, token,
                           insert_mode="INSERT_ROWS"):
    url = f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/{spreadsheet_token}/values_append"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    # None → null ใน JSON เพื่อไม่ทับเซลล์ที่มีสูตร
    if len(values) > 0 and isinstance(values[0], list):
        sheet_values = [
            [None if item is None else str(item) for item in row]
            for row in values
        ]
        max_cols = max(len(row) for row in sheet_values)
        end_col = chr(ord(start_col.upper()) + max_cols - 1)
        range_str = f"{sheet_id}!{start_col}:{end_col}"
    else:
        sheet_values = [[None if v is None else str(v)] for v in values]
        range_str = f"{sheet_id}!{start_col}:{start_col}"

    payload = {"valueRange": {"range": range_str, "values": sheet_values}}
    params = {"insertDataOption": insert_mode}

    try:
        res = requests.post(url, headers=headers, json=payload, params=params).json()
        return res
    except Exception as e:
        return {"code": -1, "msg": str(e)}


# ---------------------------------------------------------
# เขียนข้อมูลลง ยิงส่ง - ITCBI โดยไม่แตะคอลัมน์ G เลย
# (ค้นหาแถวว่าง → PUT เฉพาะ A, F, H)
# ---------------------------------------------------------
def write_song_sheet(spreadsheet_token, sheet_id, awb_list, branch_codes, timestamp, token, filename=""):
    api_headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }

    # อ่านครั้งเดียว A2:H5000 — ใช้หา next_row และนับยอดวันนี้พร้อมกัน
    read_url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{spreadsheet_token}/values/{sheet_id}!A2:H5000"
    )
    tz = timezone(timedelta(hours=7))
    today_date = datetime.now(tz).strftime("%d/%m/%Y")
    next_row = 2
    today_count_before = 0
    try:
        raw = (
            requests.get(read_url, headers={"Authorization": f"Bearer {token}"})
            .json()
            .get("data", {}).get("valueRange", {}).get("values") or []
        )
        last_idx = -1
        for i, row in enumerate(raw):
            if row and row[0] not in (None, "", "None"):
                last_idx = i
                time_val = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                if today_date in time_val:
                    today_count_before += 1
        if last_idx >= 0:
            next_row = last_idx + 3
    except Exception:
        pass

    n = len(awb_list)
    if n == 0:
        return {"code": 0}, today_count_before
    end_row = next_row + n - 1

    # Batch PUT: รวม A, F, H, I เป็น 1 API call เดียว
    f_values = [
        [branch_codes[i] if branch_codes and i < len(branch_codes) else ""]
        for i in range(n)
    ]
    value_ranges = [
        {"range": f"{sheet_id}!A{next_row}:A{end_row}", "values": [[awb] for awb in awb_list]},
        {"range": f"{sheet_id}!F{next_row}:F{end_row}", "values": f_values},
        {"range": f"{sheet_id}!H{next_row}:H{end_row}", "values": [[timestamp]] * n},
    ]
    if filename:
        value_ranges.append({"range": f"{sheet_id}!I{next_row}:I{end_row}", "values": [[filename]] * n})

    batch_url = (
        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
        f"{spreadsheet_token}/values_batch_update"
    )
    last_err = ""
    for attempt in range(3):
        resp = None
        try:
            resp = requests.put(batch_url, headers=api_headers, json={"valueRanges": value_ranges})
            raw_text = resp.text.strip()
            # ถ้า response ไม่ใช่ JSON (เช่น "404 page not found") → ล้าง stoken cache แล้ว retry
            if not raw_text.startswith('{'):
                with _api_cache_lock:
                    _api_cache["stoken"] = None
                new_stoken, _ = get_spreadsheet_token(token)
                if new_stoken:
                    batch_url = (
                        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
                        f"{new_stoken}/values_batch_update"
                    )
                last_err = f"non-JSON response: {raw_text[:80]}"
                time.sleep(1)
                continue
            res = resp.json()
            if res.get("code") == 0:
                return {"code": 0}, today_count_before + n
            last_err = res.get("msg", "batch update failed")
        except Exception as e:
            last_err = f"{str(e)} | raw={resp.text[:100] if resp is not None else 'no response'}"
        time.sleep(1)
    return {"code": -1, "msg": last_err}, today_count_before


# ---------------------------------------------------------
# ฟังก์ชันล้างข้อมูลทุกแถวใน Sheet (ลบแถวจริง ไม่ใช่แค่ล้าง)
# ---------------------------------------------------------
def clear_all_sheet_columns(spreadsheet_token, sheet_ids_dict, token):
    auth_headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json"
    }
    errors = []

    for sheet_name, sheet_id in sheet_ids_dict.items():
        # 1. อ่านคอลัมน์ A เพื่อหาแถวสุดท้ายที่มีข้อมูล
        read_url = (
            f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
            f"{spreadsheet_token}/values/{sheet_id}!A2:A5000"
        )
        last_data_row = 0
        try:
            read_res = requests.get(
                read_url, headers={"Authorization": f"Bearer {token}"}
            ).json()
            raw = read_res.get("data", {}).get("valueRange", {}).get("values") or []
            for i, row in enumerate(raw):
                if row and row[0] not in (None, "", "None"):
                    last_data_row = i + 1  # relative index (1-based from row 2)
        except Exception:
            last_data_row = 0

        if last_data_row == 0:
            continue

        # 2. ลบแถวด้วย DELETE /dimension_range
        # startIndex=1 = row 2 (0-based), endIndex=last_data_row+1 (exclusive)
        delete_url = (
            f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
            f"{spreadsheet_token}/dimension_range"
        )
        payload = {
            "dimension": {
                "sheetId": sheet_id,
                "majorDimension": "ROWS",
                "startIndex": 2,               # 1-based: row 2 = แถวข้อมูลแรก (ข้ามหัวข้อ)
                "endIndex": last_data_row + 2  # exclusive, 1-based
            }
        }
        try:
            res = requests.delete(delete_url, headers=auth_headers, json=payload)
            try:
                data = res.json()
                if data.get("code", -1) != 0:
                    errors.append(
                        f"{sheet_name}: code={data.get('code')} msg={data.get('msg', '?')}"
                    )
            except ValueError:
                if res.status_code not in (200, 204):
                    errors.append(f"{sheet_name}: HTTP {res.status_code}")
        except Exception as e:
            errors.append(f"{sheet_name}: {str(e)}")

    if errors:
        return False, " | ".join(errors)
    return True, ""


# ---------------------------------------------------------
# ฟังก์ชันตอบกลับข้อความในแชท
# ---------------------------------------------------------
def reply_message(message_id, text, token):
    """ส่งข้อความตอบกลับในแชท Feishu"""
    url = f"https://open.feishu.cn/open-apis/im/v1/messages/{message_id}/reply"
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json; charset=utf-8"
    }
    payload = {
        "msg_type": "text",
        "content": json.dumps({"text": text})
    }
    requests.post(url, headers=headers, json=payload)


# ---------------------------------------------------------
# ส่ง Interactive Card พร้อมปุ่มเมนู
# ---------------------------------------------------------
def send_menu_card(message_id, token):
    card = {
        "config": {"wide_screen_mode": True},
        "header": {
            "title": {"tag": "plain_text", "content": "🤖 JIRAYUTBOT เมนูหลัก"},
            "template": "blue"
        },
        "elements": [
            {
                "tag": "action",
                "actions": [
                    {
                        "tag": "button",
                        "text": {"tag": "plain_text", "content": "📊 สรุปยอดวันนี้"},
                        "type": "primary",
                        "value": {"command": "สรุปยอด"}
                    },
                    {
                        "tag": "button",
                        "text": {"tag": "plain_text", "content": "📅 สรุปยอดเดือนนี้"},
                        "type": "default",
                        "value": {"command": "สรุปเดือน"}
                    },
                    {
                        "tag": "button",
                        "text": {"tag": "plain_text", "content": "🗑️ ลบข้อมูลทั้งหมด"},
                        "type": "danger",
                        "value": {"command": "ลบข้อมูล"}
                    }
                ]
            }
        ]
    }
    requests.post(
        f"https://open.feishu.cn/open-apis/im/v1/messages/{message_id}/reply",
        headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json; charset=utf-8"},
        json={"msg_type": "interactive", "content": json.dumps(card)}
    )


# ---------------------------------------------------------
# ฟังก์ชันจัดการ Event แบบ Asynchronous
# ---------------------------------------------------------
def process_event(data):
    """ประมวลผล Event ใน Background Thread เพื่อป้องกัน Feishu Timeout"""
    if "header" not in data or "event" not in data:
        return

    event_type = data["header"].get("event_type")

    if event_type == "im.message.receive_v1":
        message = data["event"].get("message", {})
        message_type = message.get("message_type")
        message_id = message.get("message_id")

        token = get_tenant_access_token()
        awb_list = []
        branch_codes = []
        is_valid_input = False
        file_name = ""

        # ---------------------------------------------
        # 1. กรณีเป็นไฟล์
        # ---------------------------------------------
        if message_type == "file":
            content = json.loads(message.get("content", "{}"))
            file_key = content.get("file_key")
            file_name = content.get("file_name", "")

            if file_name.endswith(".xlsx"):
                is_valid_input = True
                reply_message(message_id, f"⏳ กำลังประมวลผลไฟล์ '{file_name}'...", token)

                file_bytes = download_feishu_file(message_id, file_key, token)
                if file_bytes:
                    awb_list, branch_codes = extract_data_from_excel(file_bytes)

                    # ดึงรหัสสาขาจากชื่อไฟล์ (ถ้าใน Excel ไม่มีรหัสสาขา)
                    if not branch_codes:
                        filename_branch_match = re.search(r'(?<!\d)(\d{6})(?!\d)', file_name)
                        if filename_branch_match:
                            branch_code_from_name = filename_branch_match.group(1)
                            if awb_list:
                                branch_codes = [branch_code_from_name] * len(awb_list)
                            else:
                                branch_codes = [branch_code_from_name]
                else:
                    reply_message(message_id, "❌ ดาวน์โหลดไฟล์ไม่สำเร็จ กรุณาลองใหม่อีกครั้ง", token)
                    return
            else:
                reply_message(message_id, f"⚠️ รองรับเฉพาะไฟล์นามสกุล .xlsx เท่านั้นครับ", token)
                return

        # ---------------------------------------------
        # 2. กรณีเป็นรูปภาพ — resize ให้พอดีแล้วตอบกลับ
        # ---------------------------------------------
        elif message_type == "image":
            from PIL import Image as PILImage

            content = json.loads(message.get("content", "{}"))
            image_key = content.get("image_key")
            if not image_key:
                return

            # ดาวน์โหลดภาพต้นฉบับ
            img_url = (
                f"https://open.feishu.cn/open-apis/im/v1/messages/"
                f"{message_id}/resources/{image_key}?type=image"
            )
            img_res = requests.get(img_url, headers={"Authorization": f"Bearer {token}"})
            if img_res.status_code != 200:
                return

            try:
                img = PILImage.open(BytesIO(img_res.content))
                w, h = img.size
                MAX_DIM = 500

                if max(w, h) <= MAX_DIM:
                    return  # ภาพเล็กพอแล้ว ไม่ต้องทำอะไร

                # คำนวณขนาดใหม่ คงสัดส่วนเดิม
                scale = MAX_DIM / max(w, h)
                new_size = (int(w * scale), int(h * scale))
                img = img.resize(new_size, PILImage.LANCZOS)

                # แปลงเป็น JPEG bytes
                if img.mode in ("RGBA", "P"):
                    img = img.convert("RGB")
                out_buf = BytesIO()
                img.save(out_buf, format="JPEG", quality=88)
                out_buf.seek(0)
                resized_bytes = out_buf.getvalue()
            except Exception:
                return

            # อัปโหลดภาพที่ resize แล้วขึ้น Feishu
            upload_res = requests.post(
                "https://open.feishu.cn/open-apis/im/v1/images",
                headers={"Authorization": f"Bearer {token}"},
                data={"image_type": "message"},
                files={"image": ("resized.jpg", resized_bytes, "image/jpeg")}
            ).json()

            new_image_key = upload_res.get("data", {}).get("image_key")
            if not new_image_key:
                return

            # ตอบกลับด้วยภาพขนาดพอดี
            requests.post(
                f"https://open.feishu.cn/open-apis/im/v1/messages/{message_id}/reply",
                headers={
                    "Authorization": f"Bearer {token}",
                    "Content-Type": "application/json; charset=utf-8"
                },
                json={
                    "msg_type": "image",
                    "content": json.dumps({"image_key": new_image_key})
                }
            )
            return

        # ---------------------------------------------
        # 3. กรณีเป็นข้อความปกติ (พิมพ์เข้ามาในแชท)
        # ---------------------------------------------
        elif message_type == "text":
            content = json.loads(message.get("content", "{}"))
            text = content.get("text", "")

            # ลบ @mention ออก (กรณีใช้ในกลุ่ม เช่น "@JIRAYUTBOT สรุปยอด" → "สรุปยอด")
            text = re.sub(r'@\S+', '', text).strip()

            # ถ้าข้อความว่าง (แค่ @mention) หรือพิมพ์ "เมนู" → แสดงการ์ดปุ่ม
            if not text or text.strip() == "เมนู":
                send_menu_card(message_id, token)
                return

            # --- ฟีเจอร์ "clear <filename>" (ลบเฉพาะไฟล์ที่ระบุ) ---
            clean_text = text.lower()
            clear_file_match = re.match(r'^clear\s+(\S+)', clean_text.strip())
            if clear_file_match:
                target_filename = clear_file_match.group(1).strip()
                try:
                    _spreadsheet_token, _error = get_spreadsheet_token(token)
                    if not _spreadsheet_token:
                        reply_message(message_id, f"❌ ไม่สามารถเข้าถึง Spreadsheet ได้: {_error}", token)
                        return

                    _sheet_ids = get_sheet_ids(_spreadsheet_token, token)
                    total_deleted = 0

                    for sname, sid in _sheet_ids.items():
                        # ยิงส่ง = column I (idx 8), ยิงถึง = column G (idx 6)
                        col_idx = 8 if sname == BRANCH_CODE_SHEET else 6
                        deleted, _ = delete_rows_by_filename(_spreadsheet_token, sid, col_idx, target_filename, token)
                        total_deleted += deleted

                    if total_deleted > 0:
                        reply_message(
                            message_id,
                            f"🗑️ ลบข้อมูลจากไฟล์ '{target_filename}' เสร็จสิ้น\n"
                            f"ลบทั้งหมด {total_deleted:,} รายการ",
                            token
                        )
                    else:
                        reply_message(
                            message_id,
                            f"⚠️ ไม่พบข้อมูลจากไฟล์ '{target_filename}' ในระบบครับ\n"
                            "(ฟีเจอร์นี้รองรับเฉพาะไฟล์ที่อัปโหลดหลังจากอัปเดตนี้)",
                            token
                        )
                except Exception as e:
                    reply_message(message_id, f"❌ เกิดข้อผิดพลาด: {str(e)}", token)
                return

            # --- ฟีเจอร์ "ลบข้อมูล" (Clear/Reset) ---
            if any(cmd in clean_text for cmd in ["clear", "ลบข้อมูล", "reset"]):
                try:
                    _spreadsheet_token, _error = get_spreadsheet_token(token)
                    if not _spreadsheet_token:
                        reply_message(message_id, f"❌ ไม่สามารถเข้าถึง Spreadsheet ได้: {_error}", token)
                        return

                    _sheet_ids = get_sheet_ids(_spreadsheet_token, token)
                    if not _sheet_ids:
                        reply_message(message_id, "❌ ไม่พบ Sheet ปลายทาง", token)
                        return

                    ok, err = clear_all_sheet_columns(_spreadsheet_token, _sheet_ids, token)

                    # รีเซ็ตตัวนับรายวัน
                    app_state["current_date"] = ""
                    app_state["daily_count"] = 0
                    app_state["branch_summary"] = {}
                    save_state()

                    if ok:
                        # เขียนสูตร XLOOKUP กลับที่ G2 ของ ยิงส่ง - ITCBI
                        song_id = _sheet_ids.get(BRANCH_CODE_SHEET)
                        if song_id:
                            try:
                                xlookup = (
                                    "=XLOOKUP(\"*\"&VALUE(F2)&\"*\","
                                    "'รอบรถ'!$D$2:$D$300,"
                                    "'รอบรถ'!$C$2:$C$300,"
                                    "\"ไม่พบรอบรถ\",2,1)"
                                )
                                requests.put(
                                    f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
                                    f"{_spreadsheet_token}/values",
                                    headers={
                                        "Authorization": f"Bearer {token}",
                                        "Content-Type": "application/json"
                                    },
                                    json={"valueRange": {"range": f"{song_id}!G2", "values": [[xlookup]]}}
                                )
                            except Exception:
                                pass
                        msg = "🗑️ เครียข้อมูล ยิงถึง-ยิงส่ง เรียบร้อยแล้ว\n📝 เขียนสูตร XLOOKUP กลับเรียบร้อยแล้ว"
                    else:
                        msg = f"❌ ล้างข้อมูลไม่สำเร็จ: {err}\nรีเซ็ตยอดรายวันเป็น 0 แล้ว"

                    reply_message(message_id, msg, token)
                    send_menu_card(message_id, token)

                except Exception as e:
                    reply_message(message_id, f"❌ เกิดข้อผิดพลาดในระบบลบข้อมูล: {str(e)}", token)

                return

            # --- ฟีเจอร์ "สรุปเดือน" (สรุปยอดสะสมทั้งเดือนจาก log sheet) ---
            if "สรุปเดือน" in text:
                tz = timezone(timedelta(hours=7))
                now = datetime.now(tz)
                month_str = now.strftime("%m/%Y")

                stoken, _err = get_spreadsheet_token(token)
                if not stoken:
                    reply_message(message_id, f"❌ ไม่สามารถเข้าถึง Sheet ได้: {_err}", token)
                    return

                sid = get_summary_sheet_id(stoken, token)
                if not sid:
                    reply_message(
                        message_id,
                        f"❌ ไม่พบ Sheet '{SUMMARY_LOG_SHEET}'\n"
                        "กรุณาสร้าง Sheet ชื่อนี้ใน Spreadsheet ก่อนครับ",
                        token
                    )
                    return

                read_url = (
                    f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
                    f"{stoken}/values/{sid}!A2:C5000"
                )
                try:
                    raw = (
                        requests.get(read_url, headers={"Authorization": f"Bearer {token}"})
                        .json()
                        .get("data", {}).get("valueRange", {}).get("values") or []
                    )
                except Exception:
                    raw = []

                monthly_branch = {}
                for row in raw:
                    if not row or not row[0]:
                        continue
                    parts = str(row[0]).strip().split("/")
                    if len(parts) == 3 and "/".join(parts[1:]) == month_str:
                        br = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                        try:
                            cnt = int(float(str(row[2]))) if len(row) > 2 and row[2] else 0
                        except (ValueError, TypeError):
                            cnt = 0
                        if br:
                            monthly_branch[br] = monthly_branch.get(br, 0) + cnt

                thai_months = ["", "ม.ค.", "ก.พ.", "มี.ค.", "เม.ย.", "พ.ค.", "มิ.ย.",
                               "ก.ค.", "ส.ค.", "ก.ย.", "ต.ค.", "พ.ย.", "ธ.ค."]
                month_num = int(now.strftime("%m"))
                year_thai = int(now.strftime("%Y")) + 543
                month_display = f"{thai_months[month_num]} {year_thai}"

                msg = f"📅 สรุปยอดเดือน {month_display}\n━━━━━━━━━━━━━━━━━━━━\n"
                if not monthly_branch:
                    msg += "ยังไม่มีข้อมูลในเดือนนี้ครับ"
                else:
                    total = 0
                    for br in sorted(monthly_branch.keys()):
                        msg += f"🏢 สาขา {br} : {monthly_branch[br]:,} รายการ\n"
                        total += monthly_branch[br]
                    msg += "━━━━━━━━━━━━━━━━━━━━\n"
                    msg += f"รวมทั้งเดือน: {total:,} รายการ"

                reply_message(message_id, msg, token)
                send_menu_card(message_id, token)
                return

            # --- ฟีเจอร์ "สรุปยอด" (อ่านจาก Sheet โดยตรง กรองเฉพาะวันนี้) ---
            if "สรุปยอด" in text:
                tz = timezone(timedelta(hours=7))
                today_str = datetime.now(tz).strftime("%d/%m/%Y")

                stoken, _err = get_spreadsheet_token(token)
                if not stoken:
                    reply_message(message_id, f"❌ ไม่สามารถเข้าถึง Sheet ได้: {_err}", token)
                    return

                sids = get_sheet_ids(stoken, token)
                song_id = sids.get(BRANCH_CODE_SHEET)

                total_count = 0
                branch_summary = {}

                if song_id:
                    read_url = (
                        f"https://open.feishu.cn/open-apis/sheets/v2/spreadsheets/"
                        f"{stoken}/values/{song_id}!A2:H5000"
                    )
                    try:
                        raw = (
                            requests.get(read_url, headers={"Authorization": f"Bearer {token}"})
                            .json()
                            .get("data", {}).get("valueRange", {}).get("values") or []
                        )
                        for row in raw:
                            if not row or not row[0] or str(row[0]).strip() in ("", "None"):
                                continue
                            # กรองเฉพาะแถวของวันนี้ (H = index 7)
                            time_val = str(row[7]).strip() if len(row) > 7 and row[7] else ""
                            if today_str not in time_val:
                                continue
                            total_count += 1
                            # F = index 5 (รหัสสาขา)
                            br = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                            if br and br not in ("", "None"):
                                branch_summary[br] = branch_summary.get(br, 0) + 1
                    except Exception:
                        pass

                summary_text = f"📊 สรุปยอดวันที่ {today_str}\n"
                summary_text += "━━━━━━━━━━━━━━━━━━━━\n"

                if total_count == 0:
                    summary_text += "ยังไม่มีข้อมูลในระบบครับ"
                else:
                    if branch_summary:
                        for br in sorted(branch_summary.keys()):
                            summary_text += f"🏢 สาขา {br} : {branch_summary[br]} รายการ\n"
                    else:
                        summary_text += "ไม่มีข้อมูลรหัสสาขา\n"
                    summary_text += "━━━━━━━━━━━━━━━━━━━━\n"
                    summary_text += f"รวมทั้งหมด: {total_count} รายการ"

                reply_message(message_id, summary_text, token)

                # บันทึก log รายวันลง Sheet สรุปยอดยิงส่ง (ถ้ามีข้อมูล)
                if total_count > 0 and branch_summary and stoken:
                    log_sid = get_summary_sheet_id(stoken, token)
                    if log_sid:
                        save_daily_log(today_str, branch_summary, stoken, log_sid, token)

                send_menu_card(message_id, token)
                return

            # ตรวจพบลูกน้ำ (,) ให้แจ้ง Error และหยุดการทำงานทันที
            if "," in text:
                reply_message(
                    message_id,
                    f"⚠️ ตรวจพบเครื่องหมาย ( , ) ในข้อความ\nERROR กรุณาลบ , ออกแล้วส่งใหม่ครับ",
                    token
                )
                return

            # ประมวลผลทีละบรรทัด ป้องกันบรรทัดเหลื่อมกัน
            lines = text.split("\n")
            current_branch = ""
            
            for line in lines:
                line_awbs = []
                # ค้นหาเลข 12 หลัก
                line_awbs.extend(re.findall(r'(?<!\d)\d{12}(?!\d)', line))
                # ค้นหารหัส B
                line_awbs.extend(re.findall(r'\bB\d+\b', line))
                # ค้นหารหัสสาขา 6 หลัก
                line_branches = re.findall(r'(?<!\d)\d{6}(?!\d)', line)

                if not line_awbs and not line_branches:
                    continue

                # ลบ AWB ซ้ำในบรรทัดเดียวกัน (แต่เก็บลำดับเดิม)
                seen = set()
                line_awbs = [x for x in line_awbs if not (x in seen or seen.add(x))]

                if line_branches:
                    # ถ้าเจอสาขา ให้จำสาขาล่าสุดเอาไว้เผื่อบรรทัดถัดไปไม่มี
                    current_branch = line_branches[-1]
                elif line_awbs:
                    # มี "ยิงไป/ยิงถึง" แต่ไม่มีรหัส 6 หลัก = ชื่อสาขา → เว้นว่าง
                    if "ยิงไป" in line or "ยิงถึง" in line:
                        line_branches = []
                        current_branch = ""
                    elif current_branch:
                        # ไม่มีคำระบุทิศทาง → สืบทอด current_branch จากบรรทัดก่อน
                        line_branches = [current_branch] * len(line_awbs)

                # ถ้าระบุรหัสสาขามา 1 ตัว แต่มีหลาย AWB ในบรรทัดนี้ ให้เบิ้ลรหัสสาขาให้เท่ากับ AWB
                if len(line_branches) == 1 and len(line_awbs) > 1:
                    line_branches = [line_branches[0]] * len(line_awbs)

                # ถ้าบรรทัดนี้มีแค่ "รหัสสาขา" (ไม่มี AWB) เราจะไม่สร้างบรรทัดว่างๆ ใน Sheet! (แค่จำไว้เฉยๆ)
                if not line_awbs:
                    continue

                # จับคู่ AWB กับ Branch เข้าด้วยกัน เพื่อป้องกันบรรทัดเหลื่อมข้ามบรรทัด
                max_len = max(len(line_awbs), len(line_branches))
                for i in range(max_len):
                    awb_list.append(line_awbs[i] if i < len(line_awbs) else "")
                    branch_codes.append(line_branches[i] if i < len(line_branches) else "")

            if awb_list or branch_codes:
                is_valid_input = True
                reply_message(message_id, f"⏳ กำลังบันทึกข้อมูลจากข้อความ...", token)
            else:
                reply_message(
                    message_id,
                    "👋 สวัสดีครับ! ส่งไฟล์ Excel (.xlsx) หรือพิมพ์เลข AWB (12 หลัก), รหัส B, รหัสสาขา (6 หลัก) มาให้ผมได้เลยครับ",
                    token
                )
                return

        # ==========================================
        # เริ่มกระบวนการเขียนลง Feishu Sheet
        # ==========================================
        if not is_valid_input:
            return

        if not awb_list and not branch_codes:
            reply_message(
                message_id,
                "⚠️ ไม่พบข้อมูล AWB (12 หลัก / รหัส B) หรือรหัสสาขา (6 หลัก) เลยครับ",
                token
            )
            return

        # ดึง Spreadsheet Token จาก Wiki
        spreadsheet_token, error_msg = get_spreadsheet_token(token)

        if not spreadsheet_token:
            reply_message(
                message_id,
                f"❌ ไม่สามารถเข้าถึง Wiki Spreadsheet ได้\n"
                f"สาเหตุ: {error_msg}\n"
                "กรุณาตรวจสอบว่า JIRAYUTBOT มีสิทธิ์เข้าถึง Wiki และ Sheet แล้ว",
                token
            )
            return

        # ดึง Sheet ID
        sheet_ids = get_sheet_ids(spreadsheet_token, token)

        if not sheet_ids:
            reply_message(
                message_id,
                "❌ ไม่พบ Sheet 'ยิงส่ง - ITCBI' หรือ 'ยิงถึง - ITCBI'\n"
                "กรุณาตรวจสอบชื่อ Sheet ใน Feishu Spreadsheet",
                token
            )
            return

        results = []
        total_awb = len(awb_list)
        tz = timezone(timedelta(hours=7))
        today_date = datetime.now(tz).strftime("%d/%m/%Y")
        today_time_str = datetime.now(tz).strftime("%d/%m/%Y %H:%M:%S")
        daily_total = total_awb  # fallback

        # เขียนทั้งสอง sheet พร้อมกัน (parallel)
        song_out = [{"code": 0}, total_awb]
        tung_out = [{"code": 0}]

        def _write_song():
            sid = sheet_ids.get(BRANCH_CODE_SHEET)
            if sid:
                res, dtotal = write_song_sheet(
                    spreadsheet_token, sid, awb_list, branch_codes, today_time_str, token, file_name
                )
                song_out[0] = res
                song_out[1] = dtotal

        def _write_tung():
            for sname, sid in sheet_ids.items():
                if sname != BRANCH_CODE_SHEET and awb_list:
                    rows = [
                        [awb, None, None, None, None, today_time_str, file_name]
                        for awb in awb_list
                    ]
                    tung_out[0] = append_to_feishu_sheet(spreadsheet_token, sid, rows, "A", token)

        t_song = threading.Thread(target=_write_song)
        t_tung = threading.Thread(target=_write_tung)
        t_song.start()
        t_tung.start()
        t_song.join()
        t_tung.join()

        daily_total = song_out[1]

        song_sid = sheet_ids.get(BRANCH_CODE_SHEET)
        if song_sid:
            if song_out[0].get("code", -1) == 0:
                results.append(f"✅ {BRANCH_CODE_SHEET}: บันทึกข้อมูลสำเร็จ")
            else:
                results.append(f"❌ {BRANCH_CODE_SHEET}: {song_out[0].get('msg', 'Unknown error')}")

        for sname in sheet_ids:
            if sname != BRANCH_CODE_SHEET:
                if tung_out[0].get("code", -1) == 0:
                    results.append(f"✅ {sname}: บันทึกข้อมูลสำเร็จ")
                else:
                    results.append(f"❌ {sname}: {tung_out[0].get('msg', 'Unknown error')}")

        # === สร้างข้อความสรุป ===
        summary = (
            f"📅 จำนวน AWB ของไฟล์นี้ : {total_awb}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
            f"📅 AWB วันที่ {today_date} รวม {daily_total}\n"
            f"━━━━━━━━━━━━━━━━━━━━\n"
        )

        errors = [msg for msg in results if "❌" in msg]
        if errors:
            summary += "\n".join(errors) + "\n━━━━━━━━━━━━━━━━━━━━\n"

        summary += "✅ ประมวลผลเสร็จสิ้น!"

        reply_message(message_id, summary, token)
        send_menu_card(message_id, token)

# ---------------------------------------------------------
# Webhook Endpoint (รับ Event จาก Feishu)
# ---------------------------------------------------------
@app.route("/webhook", methods=["POST"])
def webhook():
    data = request.json

    # ขั้นตอนที่ 1: ยืนยัน URL Challenge
    if "challenge" in data:
        return jsonify({"challenge": data["challenge"]})

    # ขั้นตอนที่ 2: ตรวจสอบ event_id ไม่ให้ประมวลผลซ้ำ
    event_id = data.get("header", {}).get("event_id")
    if event_id:
        with processed_event_lock:
            if event_id in processed_event_ids:
                return jsonify({"status": "ok"})
            processed_event_ids.add(event_id)
            # เก็บแค่ 1000 event ล่าสุด ป้องกัน memory leak
            if len(processed_event_ids) > 1000:
                oldest = next(iter(processed_event_ids))
                processed_event_ids.discard(oldest)

    # ขั้นตอนที่ 3: รันการประมวลผลใน Background Thread
    thread = threading.Thread(target=process_event, args=(data,))
    thread.start()

    # ตอบกลับ Feishu ทันทีว่า "รับทราบแล้ว" เพื่อไม่ให้ Feishu ส่งซ้ำ
    return jsonify({"status": "ok"})


# ---------------------------------------------------------
# Card Callback Endpoint (รับ Event เมื่อกดปุ่มในการ์ด)
# ---------------------------------------------------------
@app.route("/card_callback", methods=["POST"])
def card_callback():
    data = request.json or {}

    if "challenge" in data:
        return jsonify({"challenge": data["challenge"]})

    command = data.get("action", {}).get("value", {}).get("command", "")
    open_message_id = data.get("open_message_id", "")

    if command and open_message_id:
        fake_event = {
            "header": {"event_type": "im.message.receive_v1"},
            "event": {
                "message": {
                    "message_type": "text",
                    "message_id": open_message_id,
                    "content": json.dumps({"text": command})
                }
            }
        }
        threading.Thread(target=process_event, args=(fake_event,)).start()

    return jsonify({})


# ---------------------------------------------------------
# Health Check
# ---------------------------------------------------------
@app.route("/", methods=["GET"])
def health():
    return "JIRAYUTBOT is running! 🤖"


start_midnight_scheduler()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port)
